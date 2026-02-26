from pathlib import Path
from openpyxl import Workbook
from fpdf import FPDF


HEADERS = [
    "tg_id",
    "username",
    "tg_name",
    "ig_link",
    "real_name",
    "phone",
    "likes",
    "views",
    "rating",
    "created_at",
]

RATING_HEADERS = [
    "rank",
    "tg_id",
    "username",
    "tg_name",
    "ig_link",
    "likes",
    "views",
    "rating",
]


def export_users_xlsx(users, path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(HEADERS)
    for row in users:
        ws.append(list(row))
    wb.save(path)


def _safe_pdf_text(text: str) -> str:
    if text is None:
        return ""
    return text.encode("latin-1", "replace").decode("latin-1")


def export_users_pdf(users, path: Path):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 8, _safe_pdf_text("Users list"), ln=1)
    pdf.ln(2)
    for row in users:
        line = " | ".join(_safe_pdf_text(str(v)) for v in row)
        pdf.multi_cell(0, 6, line)
    pdf.output(str(path))


def export_ratings_xlsx(rows, path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ratings"
    ws.append(RATING_HEADERS)
    for row in rows:
        ws.append(list(row))
    wb.save(path)


def export_ratings_pdf(rows, path: Path):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 8, _safe_pdf_text("Rating list"), ln=1)
    pdf.ln(2)
    for row in rows:
        line = " | ".join(_safe_pdf_text(str(v)) for v in row)
        pdf.multi_cell(0, 6, line)
    pdf.output(str(path))
